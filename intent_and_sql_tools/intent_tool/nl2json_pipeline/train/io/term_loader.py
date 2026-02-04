from __future__ import annotations

import csv
import hashlib
import json
from pathlib import Path
from typing import Any, Iterable

from pydantic import BaseModel, Field

from intent_and_sql_tools.intent_tool.nl2json_pipeline.train.config import TermSourceConfig


class TermEntry(BaseModel):
    term_id: str
    name: str
    aliases: list[str] = Field(default_factory=list)
    desc: str = ""


def load_terms(config: TermSourceConfig) -> list[TermEntry]:
    path = Path(config.path)
    if not path.exists():
        raise FileNotFoundError(f"Term source not found: {path}")
    fmt = (config.format or path.suffix.lstrip(".")).lower()
    if fmt in {"jsonl", "json"}:
        records = _read_json_records(path)
    elif fmt in {"csv"}:
        records = _read_csv_records(path, config.csv_delimiter)
    elif fmt in {"xlsx", "xls"}:
        records = _read_xlsx_records(path)
    else:
        raise ValueError(f"Unsupported term source format: {fmt}")
    terms: list[TermEntry] = []
    for record in records:
        term = _build_term(record, config)
        if term:
            terms.append(term)
    return terms


def _read_json_records(path: Path) -> list[dict[str, Any]]:
    content = path.read_text(encoding="utf-8")
    if path.suffix.lower() == ".jsonl":
        records: list[dict[str, Any]] = []
        for line in content.splitlines():
            if not line.strip():
                continue
            item = json.loads(line)
            if isinstance(item, dict):
                records.append(item)
        return records
    data = json.loads(content)
    if isinstance(data, list):
        return [item for item in data if isinstance(item, dict)]
    if isinstance(data, dict):
        return [data]
    return []


def _read_csv_records(path: Path, delimiter: str) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        return [row for row in reader if isinstance(row, dict)]


def _read_xlsx_records(path: Path) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    try:
        import pandas as pd
    except Exception:
        pd = None
    if pd is not None:
        df = pd.read_excel(path)
        return df.to_dict(orient="records")
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("Reading xlsx requires pandas or openpyxl") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return records
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    for row in rows[1:]:
        item = {headers[i]: row[i] for i in range(len(headers)) if headers[i]}
        records.append(item)
    return records


def _build_term(record: dict[str, Any], config: TermSourceConfig) -> TermEntry | None:
    mapping = config.field_mapping
    term_id = _coerce_text(record.get(mapping.term_id))
    name = _coerce_text(record.get(mapping.name))
    aliases = _coerce_aliases(record.get(mapping.aliases), config.alias_delimiter)
    desc = _coerce_text(record.get(mapping.desc))
    if not name:
        return None
    if not term_id:
        term_id = _stable_id(name)
    return TermEntry(term_id=term_id, name=name, aliases=aliases, desc=desc)


def _coerce_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _coerce_aliases(value: Any, delimiter: str) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    if isinstance(value, str):
        if not value.strip():
            return []
        raw = value.strip()
        if raw.startswith("[") and raw.endswith("]"):
            try:
                data = json.loads(raw)
                if isinstance(data, list):
                    return [str(v).strip() for v in data if str(v).strip()]
            except json.JSONDecodeError:
                pass
        parts = [p.strip() for p in value.split(delimiter)]
        return [p for p in parts if p]
    return [str(value).strip()]


def _stable_id(text: str) -> str:
    digest = hashlib.sha1(text.encode("utf-8")).hexdigest()[:10]
    return f"term_{digest}"


def iter_batches(items: Iterable[TermEntry], batch_size: int) -> Iterable[list[TermEntry]]:
    batch: list[TermEntry] = []
    for item in items:
        batch.append(item)
        if len(batch) >= batch_size:
            yield batch
            batch = []
    if batch:
        yield batch
