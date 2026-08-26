"""Local, content-addressed Source → Profile → Clean → Golden pipeline."""

from __future__ import annotations

import csv
import hashlib
import json
import os
import sqlite3
import uuid
from dataclasses import dataclass
from pathlib import Path

from .adapters import (
    LocalSourceAdapter,
    _bounded_integer,
    _infer_mapping_fields,
    _read_json_rows,
    _read_parquet,
    _read_pdf_rows,
    _read_text_document,
)
from .models import (
    ArtifactRef,
    AssetOwner,
    AssetPermission,
    CleaningOperation,
    CleaningRecipeRecord,
    CleanRunRecord,
    ConnectionInstance,
    GoldenAssetRevisionRecord,
    GoldenLineage,
    ProfileField,
    ProfileRunRecord,
    SourceRevisionRecord,
    SourceType,
)


class ContentAddressedArtifactStore:
    """A bounded local artifact store addressed only by verified SHA-256."""

    def __init__(self, root: Path) -> None:
        self.root = root.resolve()
        self.root.mkdir(parents=True, exist_ok=True)

    def put(self, content: bytes, media_type: str) -> ArtifactRef:
        digest = hashlib.sha256(content).hexdigest()
        target = self.root / digest[:2] / digest
        target.parent.mkdir(parents=True, exist_ok=True)
        if not target.exists():
            temporary = target.with_name(f".{digest}.{uuid.uuid4().hex}.tmp")
            temporary.write_bytes(content)
            os.replace(temporary, target)
        return ArtifactRef(
            uri=f"artifact://sha256/{digest}",
            sha256=digest,
            media_type=media_type,
            bytes=len(content),
        )

    def read(self, ref: ArtifactRef) -> bytes:
        prefix = "artifact://sha256/"
        if not ref.uri.startswith(prefix):
            raise ValueError("artifact reference is not owned by this store")
        digest = ref.uri.removeprefix(prefix)
        if digest != ref.sha256 or len(digest) != 64:
            raise ValueError("artifact reference digest mismatch")
        content = (self.root / digest[:2] / digest).read_bytes()
        if hashlib.sha256(content).hexdigest() != digest:
            raise ValueError("artifact content digest mismatch")
        return content


@dataclass(frozen=True)
class MaterializedSource:
    source_type: SourceType
    source_locator: str
    raw_content: bytes
    rows: list[dict[str, object]]
    fields: list[tuple[str, str, bool]]
    media_type: str
    adapter_run_id: str | None = None
    checkpoint: dict[str, str] | None = None


@dataclass(frozen=True)
class LifecycleRecords:
    source: SourceRevisionRecord
    profile: ProfileRunRecord
    recipe: CleaningRecipeRecord
    clean: CleanRunRecord
    golden: GoldenAssetRevisionRecord


class LocalLifecycle:
    def __init__(
        self, *, source_root: Path, artifact_store: ContentAddressedArtifactStore
    ):
        self.source_root = source_root.resolve()
        self.artifact_store = artifact_store

    def build(
        self,
        *,
        connection: ConnectionInstance,
        resource_id: str,
        operations: list[CleaningOperation],
        recipe_version: int,
        golden_revision: int,
        principal_id: str,
        trace_id: str,
        timestamp: str,
        materialized: MaterializedSource | None = None,
        tool_arguments: dict[str, object] | None = None,
    ) -> LifecycleRecords:
        materialized = materialized or self.materialize(connection, resource_id)
        raw_ref = self.artifact_store.put(
            materialized.raw_content, materialized.media_type
        )
        schema_payload = {
            "sourceType": materialized.source_type,
            "fields": [
                {"name": name, "type": data_type, "nullable": nullable}
                for name, data_type, nullable in materialized.fields
            ],
        }
        schema_digest = _digest_json(schema_payload)
        source_id = (
            "source-"
            + _digest_json(
                {
                    "connectionId": connection.id,
                    "resourceId": resource_id,
                    "sourceDigest": raw_ref.sha256,
                    "schemaDigest": schema_digest,
                }
            )[:24]
        )
        source = SourceRevisionRecord(
            id=source_id,
            workspace_id=connection.workspace_id,
            connection_id=connection.id,
            resource_id=resource_id,
            source_type=materialized.source_type,
            content_ref=raw_ref,
            source_digest=raw_ref.sha256,
            schema_digest=schema_digest,
            source_locator=materialized.source_locator,
            permission_version=1,
            checkpoint=materialized.checkpoint or {},
            created_at=timestamp,
            trace_id=trace_id,
        )

        sensitive = [
            name for name, _, _ in materialized.fields if _is_sensitive_field(name)
        ]
        profile_fields = [
            ProfileField(
                name=name,
                data_type=data_type,
                nullable=nullable,
                null_count=sum(
                    row.get(name) in (None, "") for row in materialized.rows
                ),
                distinct_count=len(
                    {
                        _hashable(row.get(name))
                        for row in materialized.rows
                        if row.get(name) not in (None, "")
                    }
                ),
                sensitive=name in sensitive,
            )
            for name, data_type, nullable in materialized.fields
        ]
        cell_count = len(materialized.rows) * max(len(profile_fields), 1)
        nonempty = sum(
            value not in (None, "")
            for row in materialized.rows
            for value in row.values()
        )
        quality_score = nonempty / cell_count if cell_count else 1.0
        safe_sample = [_redact_row(row, sensitive) for row in materialized.rows[:100]]
        sample_ref = self.artifact_store.put(
            _json_bytes(safe_sample), "application/json"
        )
        report = {
            "rowCount": len(materialized.rows),
            "fields": [
                field.model_dump(mode="json", by_alias=True) for field in profile_fields
            ],
            "qualityScore": quality_score,
            "sensitiveFields": sensitive,
        }
        report_ref = self.artifact_store.put(_json_bytes(report), "application/json")
        profile_id = (
            f"profile-{source.id.removeprefix('source-')}-"
            f"{hashlib.sha256(trace_id.encode()).hexdigest()[:8]}"
        )
        profile = ProfileRunRecord(
            id=profile_id,
            source_revision_id=source.id,
            status="succeeded",
            row_count=len(materialized.rows),
            fields=profile_fields,
            quality_score=quality_score,
            sensitive_fields=sensitive,
            report_ref=report_ref,
            sample_ref=sample_ref,
            started_at=timestamp,
            finished_at=timestamp,
            trace_id=trace_id,
        )

        asset_id = (
            "golden-"
            + hashlib.sha256(f"{connection.id}:{resource_id}".encode()).hexdigest()[:24]
        )
        normalized_operations = self._validate_operations(operations)
        recipe_digest = _digest_json(
            {
                "assetId": asset_id,
                "version": recipe_version,
                "sourceRevisionId": source.id,
                "operations": normalized_operations,
            }
        )
        recipe = CleaningRecipeRecord(
            id=f"recipe-{asset_id.removeprefix('golden-')}",
            asset_id=asset_id,
            version=recipe_version,
            source_revision_id=source.id,
            operations=normalized_operations,
            recipe_digest=recipe_digest,
            created_at=timestamp,
        )
        cleaned_rows = _clean_rows(materialized.rows, normalized_operations, sensitive)
        output = b"".join(
            json.dumps(row, ensure_ascii=False, sort_keys=True).encode() + b"\n"
            for row in cleaned_rows
        )
        output_ref = self.artifact_store.put(output, "application/x-ndjson")
        quality_report_ref = self.artifact_store.put(
            _json_bytes(
                {
                    "sourceRows": len(materialized.rows),
                    "outputRows": len(cleaned_rows),
                    "qualityScore": quality_score,
                    "schemaDigest": schema_digest,
                }
            ),
            "application/json",
        )
        clean_id = (
            f"clean-{asset_id.removeprefix('golden-')}-v{recipe_version}-"
            f"{output_ref.sha256[:8]}"
        )
        clean = CleanRunRecord(
            id=clean_id,
            source_revision_id=source.id,
            recipe_id=recipe.id,
            status="succeeded",
            output_ref=output_ref,
            quality_report_ref=quality_report_ref,
            started_at=timestamp,
            finished_at=timestamp,
            trace_id=trace_id,
        )
        lineage_digest = _digest_json(
            {
                "connectionId": connection.id,
                "resourceId": resource_id,
                "sourceRevisionId": source.id,
                "profileRunId": profile.id,
                "recipeId": recipe.id,
                "recipeVersion": recipe.version,
                "cleanRunId": clean.id,
                "contentDigest": source.source_digest,
                "correlationId": trace_id,
                "adapterRunId": materialized.adapter_run_id,
                "checkpoint": materialized.checkpoint or {},
                "outputDigest": output_ref.sha256,
                "toolArguments": tool_arguments or {},
            }
        )
        lineage = GoldenLineage(
            connection_id=connection.id,
            resource_id=resource_id,
            source_revision_id=source.id,
            profile_run_id=profile.id,
            recipe_id=recipe.id,
            recipe_version=recipe.version,
            clean_run_id=clean.id,
            content_digest=source.source_digest,
            correlation_id=trace_id,
            adapter_run_id=materialized.adapter_run_id,
            checkpoint=materialized.checkpoint or {},
            lineage_digest=lineage_digest,
            tool_arguments=tool_arguments or {},
        )
        golden = GoldenAssetRevisionRecord(
            id=(f"{asset_id}-r{golden_revision}-{output_ref.sha256[:8]}"),
            asset_id=asset_id,
            revision=golden_revision,
            asset_kind=(
                "knowledge"
                if materialized.source_type in {"markdown", "pdf"}
                else "dataset"
            ),
            schema_digest=schema_digest,
            storage_ref=output_ref,
            owner=AssetOwner(
                workspace_id=connection.workspace_id,
                principal_id=principal_id,
            ),
            permissions=AssetPermission(
                workspace_id=connection.workspace_id,
                scope=connection.scope,
                can_read=True,
                can_write=True,
                inherited_from_connection_id=connection.id,
                version=1,
            ),
            lineage=lineage,
            quality_score=quality_score,
            freshness_at=timestamp,
            data_as_of=_data_as_of(cleaned_rows, timestamp),
            trace_id=trace_id,
        )
        return LifecycleRecords(
            source=source,
            profile=profile,
            recipe=recipe,
            clean=clean,
            golden=golden,
        )

    @staticmethod
    def materialized_mcp(
        *,
        tool_name: str,
        rows: list[dict[str, object]],
        fields: list[tuple[str, str, bool]],
        adapter_run_id: str,
    ) -> MaterializedSource:
        raw = _json_bytes(rows)
        return MaterializedSource(
            source_type="mcp",
            source_locator=f"mcp://tool/{tool_name}",
            raw_content=raw,
            rows=rows,
            fields=fields,
            media_type="application/json",
            adapter_run_id=adapter_run_id,
        )

    def materialize(
        self, connection: ConnectionInstance, resource_id: str
    ) -> MaterializedSource:
        source_ref = connection.configuration.get("sourceRef")
        if not isinstance(source_ref, str):
            raise TypeError("persisted connection has no sourceRef")
        path = LocalSourceAdapter(root=self.source_root).resolve(
            source_ref, connection.connector_key
        )
        raw = path.read_bytes()
        if connection.connector_key == "local_file":
            text = raw.decode("utf-8")
            return MaterializedSource(
                source_type="markdown",
                source_locator=source_ref,
                raw_content=raw,
                rows=[{"text": line} for line in text.splitlines() if line.strip()],
                fields=[("text", "string", False)],
                media_type="text/markdown",
            )
        if connection.connector_key == "doc_txt":
            if path.suffix.lower() != ".pdf":
                max_chars = connection.configuration.get("maxTextChars", 2_000_000)
                if isinstance(max_chars, bool) or not isinstance(max_chars, int):
                    raise ValueError("document character limit must be an integer")
                source_type, text = _read_text_document(path, max_chars=max_chars)
                return MaterializedSource(
                    source_type=source_type,
                    source_locator=source_ref,
                    raw_content=raw,
                    rows=[{"text": line} for line in text.splitlines() if line.strip()],
                    fields=[("text", "string", False)],
                    media_type={
                        "markdown": "text/markdown",
                        "text": "text/plain",
                        "html": "text/html",
                    }[source_type],
                )
            max_chars = connection.configuration.get("maxTextChars", 2_000_000)
            if isinstance(max_chars, bool) or not isinstance(max_chars, int):
                raise ValueError("document character limit must be an integer")
            rows = _read_pdf_rows(path, max_chars=max_chars)
            return MaterializedSource(
                source_type="pdf",
                source_locator=source_ref,
                raw_content=raw,
                rows=rows,
                fields=[
                    ("page", "integer", False),
                    ("text", "string", False),
                ],
                media_type="application/pdf",
            )
        if connection.connector_key == "csv":
            text = raw.decode("utf-8")
            reader = csv.DictReader(text.splitlines())
            if not reader.fieldnames or any(not name for name in reader.fieldnames):
                raise ValueError("CSV must have a non-empty header")
            if len(set(reader.fieldnames)) != len(reader.fieldnames):
                raise ValueError("CSV column names must be unique")
            fieldnames = list(reader.fieldnames)
            raw_rows = [{key: value for key, value in row.items()} for row in reader]
            typed_rows, fields = _infer_csv(raw_rows, fieldnames)
            return MaterializedSource(
                source_type="csv",
                source_locator=source_ref,
                raw_content=raw,
                rows=typed_rows,
                fields=fields,
                media_type="text/csv",
            )
        if connection.connector_key == "json":
            max_depth = connection.configuration.get("maxDepth", 32)
            max_rows = connection.configuration.get("maxRows", 10_000)
            if (
                isinstance(max_depth, bool)
                or not isinstance(max_depth, int)
                or isinstance(max_rows, bool)
                or not isinstance(max_rows, int)
            ):
                raise ValueError("JSON limits must be integers")
            rows = _read_json_rows(
                path,
                max_depth=max_depth,
                max_rows=max_rows,
            )
            return MaterializedSource(
                source_type="json",
                source_locator=source_ref,
                raw_content=raw,
                rows=rows,
                fields=_infer_mapping_fields(rows),
                media_type="application/json",
            )
        if connection.connector_key == "parquet":
            options = connection.configuration
            rows, fields = _read_parquet(
                path,
                max_rows=_bounded_integer(
                    options,
                    "maxRows",
                    default=10_000,
                    maximum=1_000_000,
                ),
                max_columns=_bounded_integer(
                    options,
                    "maxColumns",
                    default=1_000,
                    maximum=10_000,
                ),
                max_uncompressed_bytes=_bounded_integer(
                    options,
                    "maxUncompressedBytes",
                    default=100 * 1024 * 1024,
                    maximum=1024 * 1024 * 1024,
                ),
                max_nesting_depth=_bounded_integer(
                    options,
                    "maxNestingDepth",
                    default=16,
                    maximum=64,
                ),
            )
            return MaterializedSource(
                source_type="parquet",
                source_locator=source_ref,
                raw_content=raw,
                rows=rows,
                fields=fields,
                media_type="application/vnd.apache.parquet",
            )
        if connection.connector_key == "excel":
            return self._materialize_excel(connection, resource_id, path, raw)
        if connection.connector_key != "sqlite":
            raise ValueError("connector has no local lifecycle adapter")
        resource = next(
            (
                item
                for item in connection.discovered_resources
                if item.id == resource_id and item.resource_type == "table"
            ),
            None,
        )
        if resource is None:
            raise ValueError("SQLite table was not discovered by this connection")
        quoted = resource.name.replace('"', '""')
        row_limit_value = connection.configuration.get("rowLimit", 10_000)
        if isinstance(row_limit_value, bool) or not isinstance(row_limit_value, int):
            raise TypeError("SQLite rowLimit must be an integer")
        row_limit = row_limit_value
        database = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
        database.row_factory = sqlite3.Row
        try:
            rows = [
                dict(row)
                for row in database.execute(
                    f'SELECT * FROM "{quoted}" LIMIT ?', (row_limit,)
                ).fetchall()
            ]
        finally:
            database.close()
        return MaterializedSource(
            source_type="sqlite",
            source_locator=f"{source_ref}#{resource.name}",
            raw_content=raw,
            rows=rows,
            fields=[
                (field.name, field.data_type, field.nullable)
                for field in resource.fields
            ],
            media_type="application/vnd.sqlite3",
        )

    @staticmethod
    def _materialize_excel(
        connection: ConnectionInstance,
        resource_id: str,
        path: Path,
        raw: bytes,
    ) -> MaterializedSource:
        resource = next(
            (
                item
                for item in connection.discovered_resources
                if item.id == resource_id and item.resource_type == "table"
            ),
            None,
        )
        if resource is None:
            raise ValueError("Excel sheet was not discovered by this connection")
        values: list[tuple[object, ...]]
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise ValueError(
                "Excel adapter requires the server-side openpyxl dependency"
            ) from error
        workbook = load_workbook(
            str(path), read_only=True, data_only=True, keep_links=False
        )
        try:
            sheet = workbook[resource.name]
            iterator = sheet.iter_rows(values_only=True)
            header = next(iterator, ())
            names = _validated_header(header, "Excel")
            values = [tuple(row) for row in iterator]
        finally:
            workbook.close()
        rows: list[dict[str, object]] = [
            {
                name: row[index] if index < len(row) else None
                for index, name in enumerate(names)
            }
            for row in values
            if any(value not in (None, "") for value in row)
        ]
        return MaterializedSource(
            source_type="excel",
            source_locator=f"{connection.configuration['sourceRef']}#{resource.name}",
            raw_content=raw,
            rows=rows,
            fields=[
                (field.name, field.data_type, field.nullable)
                for field in resource.fields
            ],
            media_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )

    @staticmethod
    def _validate_operations(
        operations: list[CleaningOperation],
    ) -> list[CleaningOperation]:
        allowed: set[CleaningOperation] = {
            "trim",
            "deduplicate",
            "normalize",
            "redact",
        }
        unknown = set(operations) - allowed
        if unknown:
            raise ValueError(f"unsupported cleaning operations: {sorted(unknown)}")
        return list(dict.fromkeys(operations))


def _digest_json(value: object) -> str:
    return hashlib.sha256(
        json.dumps(
            value, ensure_ascii=False, sort_keys=True, separators=(",", ":")
        ).encode()
    ).hexdigest()


def _validated_header(values: object, label: str) -> list[str]:
    if not isinstance(values, (list, tuple)) or not values:
        raise ValueError(f"{label} source must have a non-empty header row")
    names = [str(value).strip() if value is not None else "" for value in values]
    if any(not name for name in names):
        raise ValueError(f"{label} column names must be non-empty")
    if len(set(names)) != len(names):
        raise ValueError(f"{label} column names must be unique")
    return names


def _json_bytes(value: object) -> bytes:
    return json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode()


def _hashable(value: object) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def _is_sensitive_field(name: str) -> bool:
    lowered = name.casefold()
    return any(
        marker in lowered
        for marker in (
            "email",
            "phone",
            "mobile",
            "token",
            "secret",
            "password",
            "credential",
        )
    )


def _redact_row(
    row: dict[str, object], sensitive_fields: list[str]
) -> dict[str, object]:
    return {
        key: "[REDACTED]"
        if key in sensitive_fields and value not in (None, "")
        else value
        for key, value in row.items()
    }


def _infer_csv(
    rows: list[dict[str, str]], fieldnames: list[str]
) -> tuple[list[dict[str, object]], list[tuple[str, str, bool]]]:
    kinds: dict[str, str] = {}
    nullable: dict[str, bool] = {}
    for name in fieldnames:
        values = [row.get(name, "") for row in rows]
        present = [value.strip() for value in values if value and value.strip()]
        nullable[name] = len(present) != len(values)
        if present and all(_is_integer(value) for value in present):
            kinds[name] = "integer"
        elif present and all(_is_number(value) for value in present):
            kinds[name] = "number"
        else:
            kinds[name] = "string"
    typed: list[dict[str, object]] = []
    for row in rows:
        normalized: dict[str, object] = {}
        for name in fieldnames:
            value = (row.get(name) or "").strip()
            if not value:
                normalized[name] = None
            elif kinds[name] == "integer":
                normalized[name] = int(value)
            elif kinds[name] == "number":
                normalized[name] = float(value)
            else:
                normalized[name] = value
        typed.append(normalized)
    return typed, [(name, kinds[name], nullable[name]) for name in fieldnames]


def _is_integer(value: str) -> bool:
    try:
        int(value)
        return "." not in value
    except ValueError:
        return False


def _is_number(value: str) -> bool:
    try:
        float(value)
        return True
    except ValueError:
        return False


def _clean_rows(
    rows: list[dict[str, object]],
    operations: list[CleaningOperation],
    sensitive_fields: list[str],
) -> list[dict[str, object]]:
    cleaned: list[dict[str, object]] = []
    seen: set[str] = set()
    for original in rows:
        row = dict(original)
        if "trim" in operations:
            row = {
                key: value.strip() if isinstance(value, str) else value
                for key, value in row.items()
            }
        if "normalize" in operations:
            row = {
                key: value.casefold() if isinstance(value, str) else value
                for key, value in row.items()
            }
        redact_before_deduplication = "redact" in operations
        if "redact" in operations:
            row = _redact_row(row, sensitive_fields)
        marker = _hashable(row)
        if "deduplicate" in operations and marker in seen:
            continue
        seen.add(marker)
        cleaned.append(
            row if redact_before_deduplication else _redact_row(row, sensitive_fields)
        )
    return cleaned


def _data_as_of(rows: list[dict[str, object]], fallback: str) -> str:
    values = [
        str(row["dataAsOf"]) for row in rows if row.get("dataAsOf") not in (None, "")
    ]
    return max(values) if values else fallback
