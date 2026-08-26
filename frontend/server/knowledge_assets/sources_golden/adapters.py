"""Concrete local adapters and fail-closed provider contracts."""

from __future__ import annotations

import csv
import hashlib
import sqlite3
import zipfile
from collections.abc import Callable, Sequence
from pathlib import Path

from ..security import (
    validate_archive_limits,
    validate_database_limits,
    validate_mcp_tool,
    validate_read_only_sql,
)
from .local_formats import (
    _bounded_integer,
    _infer_mapping_fields,
    _read_json_rows,
    _read_parquet,
    _read_pdf_rows,
    _read_text_document,
)
from .models import (
    AdapterContract,
    CapabilityReason,
    ConnectorDefinition,
    ConnectorOperation,
    ConnectorOperationName,
    DiscoveredField,
    DiscoveredResource,
)


class LocalSourceAdapter:
    def __init__(self, *, root: Path, max_bytes: int = 10 * 1024 * 1024) -> None:
        self.root = root.resolve()
        self.max_bytes = max_bytes

    def resolve(self, source_ref: object, connector_key: str) -> Path:
        if not isinstance(source_ref, str) or not source_ref:
            raise ValueError("sourceRef is required")
        unresolved = self.root / source_ref
        if unresolved.is_symlink():
            raise ValueError("source symlinks are not allowed")
        path = unresolved.resolve()
        if path != self.root and self.root not in path.parents:
            raise ValueError("source escapes the configured workspace root")
        if not path.is_file():
            raise FileNotFoundError(path)
        if path.stat().st_size > self.max_bytes:
            raise ValueError("source exceeds the configured byte limit")
        expected = {
            "csv": {".csv"},
            "json": {".json", ".jsonl", ".ndjson"},
            "parquet": {".parquet"},
            "doc_txt": {".pdf", ".md", ".markdown", ".txt", ".html", ".htm"},
            "excel": {".xlsx"},
            "local_file": {".md", ".markdown"},
            "sqlite": {".sqlite", ".sqlite3", ".db"},
        }[connector_key]
        if path.suffix.lower() not in expected:
            raise ValueError(f"{connector_key} source has an unsupported extension")
        if connector_key == "excel":
            self._validate_xlsx_archive(path)
        if (
            connector_key in {"csv", "json", "local_file"}
            or connector_key == "doc_txt"
            and path.suffix.lower() != ".pdf"
        ) and b"\x00" in path.read_bytes():
            raise ValueError("text source contains binary NUL bytes")
        return path

    @staticmethod
    def _validate_xlsx_archive(path: Path) -> None:
        try:
            with zipfile.ZipFile(path) as archive:
                members = archive.infolist()
                validate_archive_limits(
                    compressed_bytes=sum(item.compress_size for item in members),
                    expanded_bytes=sum(item.file_size for item in members),
                    file_count=len(members),
                    member_names=[item.filename.replace("\\", "/") for item in members],
                )
        except zipfile.BadZipFile as error:
            raise ValueError("Excel workbook is not a valid XLSX archive") from error

    def validate(
        self,
        *,
        connector_key: str,
        configuration: dict[str, object],
        trace_id: str,
    ) -> tuple[Path, ConnectorOperation]:
        path = self.resolve(configuration.get("sourceRef"), connector_key)
        return path, ConnectorOperation(
            operation="validate",
            status="succeeded",
            trace_id=trace_id,
            reason=CapabilityReason(
                code="VALIDATED",
                message="The selected local source passed format, path, and size checks.",
            ),
        )

    def discover(
        self,
        *,
        connector_key: str,
        path: Path,
        trace_id: str,
        configuration: dict[str, object] | None = None,
    ) -> ConnectorOperation:
        digest = hashlib.sha256(path.read_bytes()).hexdigest()
        if connector_key == "csv":
            with path.open(newline="", encoding="utf-8") as handle:
                reader = csv.reader(handle)
                header = next(reader, [])
                row_count = sum(1 for _ in reader)
            resources = [
                DiscoveredResource(
                    id=f"file-{digest[:24]}",
                    name=path.name,
                    resource_type="file",
                    row_count=row_count,
                    fields=[
                        DiscoveredField(name=name, data_type="string")
                        for name in header
                    ],
                )
            ]
        elif connector_key == "json":
            rows = _read_json_rows(
                path,
                max_depth=_bounded_integer(
                    configuration or {}, "maxDepth", default=32, maximum=128
                ),
                max_rows=_bounded_integer(
                    configuration or {}, "maxRows", default=10_000, maximum=1_000_000
                ),
            )
            fields = _infer_mapping_fields(rows)
            resources = [
                DiscoveredResource(
                    id=f"file-{digest[:24]}",
                    name=path.name,
                    resource_type="file",
                    row_count=len(rows),
                    fields=[
                        DiscoveredField(
                            name=name,
                            data_type=data_type,
                            nullable=nullable,
                        )
                        for name, data_type, nullable in fields
                    ],
                )
            ]
        elif connector_key == "parquet":
            rows, fields = _read_parquet(
                path,
                max_rows=_bounded_integer(
                    configuration or {}, "maxRows", default=10_000, maximum=1_000_000
                ),
                max_columns=_bounded_integer(
                    configuration or {}, "maxColumns", default=1_000, maximum=10_000
                ),
                max_uncompressed_bytes=_bounded_integer(
                    configuration or {},
                    "maxUncompressedBytes",
                    default=100 * 1024 * 1024,
                    maximum=1024 * 1024 * 1024,
                ),
                max_nesting_depth=_bounded_integer(
                    configuration or {},
                    "maxNestingDepth",
                    default=16,
                    maximum=64,
                ),
            )
            resources = [
                DiscoveredResource(
                    id=f"file-{digest[:24]}",
                    name=path.name,
                    resource_type="file",
                    row_count=len(rows),
                    fields=[
                        DiscoveredField(
                            name=name,
                            data_type=data_type,
                            nullable=nullable,
                        )
                        for name, data_type, nullable in fields
                    ],
                )
            ]
        elif connector_key == "excel":
            resources = self._discover_excel(path, digest, configuration or {})
        elif connector_key in {"local_file", "doc_txt"}:
            max_chars = _bounded_integer(
                configuration or {},
                "maxTextChars",
                default=2_000_000,
                maximum=10_000_000,
            )
            if connector_key == "doc_txt" and path.suffix.lower() == ".pdf":
                _read_pdf_rows(path, max_chars=max_chars)
            elif connector_key == "doc_txt":
                _read_text_document(path, max_chars=max_chars)
            resources = [
                DiscoveredResource(
                    id=f"document-{digest[:24]}",
                    name=path.name,
                    resource_type="document",
                    fields=(
                        [
                            DiscoveredField(name="page", data_type="integer"),
                            DiscoveredField(name="text", data_type="string"),
                        ]
                        if connector_key == "doc_txt" and path.suffix.lower() == ".pdf"
                        else [DiscoveredField(name="text", data_type="string")]
                    ),
                )
            ]
        else:
            resources = self._discover_sqlite(path, digest)
        return ConnectorOperation(
            operation="discover",
            status="succeeded",
            trace_id=trace_id,
            reason=CapabilityReason(
                code="DISCOVERED",
                message="Resources were discovered from the selected local source.",
            ),
            resources=resources,
        )

    @staticmethod
    def _discover_excel(
        path: Path,
        digest: str,
        configuration: dict[str, object],
    ) -> list[DiscoveredResource]:
        sheet_allowlist = configuration.get("sheetAllowlist", [])
        if not isinstance(sheet_allowlist, list) or not all(
            isinstance(item, str) and item for item in sheet_allowlist
        ):
            raise ValueError("Excel sheetAllowlist must contain sheet names")
        resources: list[DiscoveredResource] = []
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise ValueError(
                "Excel adapter requires the server-side openpyxl dependency"
            ) from error
        workbook = load_workbook(
            str(path), read_only=True, data_only=True, keep_links=False
        )
        try:
            available = set(workbook.sheetnames)
            unknown = set(sheet_allowlist) - available
            if unknown:
                raise ValueError(f"Excel sheets were not found: {sorted(unknown)}")
            selected = sheet_allowlist or workbook.sheetnames
            for sheet_name in selected:
                sheet = workbook[sheet_name]
                rows = sheet.iter_rows(values_only=True)
                header = next(rows, ())
                names = _excel_header(header)
                values = [tuple(row) for row in rows]
                resources.append(
                    _excel_resource(
                        digest=digest,
                        sheet_name=sheet_name,
                        names=names,
                        rows=values,
                    )
                )
        finally:
            workbook.close()
        return resources

    @staticmethod
    def _discover_sqlite(path: Path, digest: str) -> list[DiscoveredResource]:
        connection: sqlite3.Connection | None = None
        try:
            connection = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
            connection.row_factory = sqlite3.Row
            tables = connection.execute(
                """
                SELECT name FROM sqlite_master
                WHERE type = 'table' AND name NOT LIKE 'sqlite_%'
                ORDER BY name
                """
            ).fetchall()
            resources: list[DiscoveredResource] = []
            for table in tables:
                table_name = str(table["name"])
                quoted = table_name.replace('"', '""')
                fields = connection.execute(f'PRAGMA table_info("{quoted}")').fetchall()
                row_count = connection.execute(
                    f'SELECT COUNT(*) AS count FROM "{quoted}"'
                ).fetchone()["count"]
                resources.append(
                    DiscoveredResource(
                        id=f"table-{digest[:16]}-{hashlib.sha256(table_name.encode()).hexdigest()[:8]}",
                        name=table_name,
                        schema_name="main",
                        resource_type="table",
                        row_count=row_count,
                        fields=[
                            DiscoveredField(
                                name=str(field["name"]),
                                data_type=str(field["type"] or "BLOB").lower(),
                                nullable=not bool(field["notnull"]),
                            )
                            for field in fields
                        ],
                    )
                )
            return resources
        except sqlite3.DatabaseError as error:
            raise ValueError("invalid or unsafe SQLite database") from error
        finally:
            if connection is not None:
                connection.close()


def _excel_header(values: object) -> list[str]:
    if not isinstance(values, (list, tuple)) or not values:
        raise ValueError("Excel sheet must have a non-empty header row")
    names = [str(value).strip() if value is not None else "" for value in values]
    if any(not name for name in names):
        raise ValueError("Excel column names must be non-empty")
    if len(set(names)) != len(names):
        raise ValueError("Excel column names must be unique")
    return names


def _excel_resource(
    *,
    digest: str,
    sheet_name: str,
    names: list[str],
    rows: Sequence[Sequence[object]],
) -> DiscoveredResource:
    return DiscoveredResource(
        id=(
            f"sheet-{digest[:16]}-{hashlib.sha256(sheet_name.encode()).hexdigest()[:8]}"
        ),
        name=sheet_name,
        schema_name="workbook",
        resource_type="table",
        row_count=len(rows),
        fields=[
            DiscoveredField(
                name=name,
                data_type=_excel_field_type(
                    [row[index] if index < len(row) else None for row in rows]
                ),
                nullable=any(
                    index >= len(row) or row[index] in (None, "") for row in rows
                ),
            )
            for index, name in enumerate(names)
        ],
    )


def _excel_field_type(values: list[object]) -> str:
    present = [value for value in values if value not in (None, "")]
    if present and all(isinstance(value, bool) for value in present):
        return "boolean"
    if present and all(
        isinstance(value, int) and not isinstance(value, bool) for value in present
    ):
        return "integer"
    if present and all(
        isinstance(value, (int, float)) and not isinstance(value, bool)
        for value in present
    ):
        return "number"
    return "string"


def blocked_operation(
    definition: ConnectorDefinition,
    *,
    operation: ConnectorOperationName,
    trace_id: str,
) -> ConnectorOperation:
    status = (
        "config_required"
        if definition.capability_state == "configurable"
        else definition.capability_state
    )
    if status == "available":
        raise ValueError("available connectors require a concrete local adapter")
    return ConnectorOperation(
        operation=operation,
        status=status,
        trace_id=trace_id,
        reason=definition.reason,
    )


def adapter_contract(definition: ConnectorDefinition) -> AdapterContract:
    key = definition.connector_key
    database = {
        "postgresql",
        "mysql",
        "oracle",
        "sqlserver",
        "clickhouse",
        "doris",
        "starrocks",
        "snowflake",
        "bigquery",
        "hive",
        "sqlite",
    }
    lark = {
        "lark_doc",
        "lark_wiki",
        "lark_drive",
        "lark_meeting",
        "lark_minutes",
        "lark_group",
        "lark_chat",
        "lark_sheet",
        "lark_base",
        "lark_mail",
    }
    web = {"rest_api", "graphql", "web_discovery", "webhook", "custom_http"}
    protocol = (
        key
        if key in database
        else "lark_openapi"
        if key in lark
        else "mcp"
        if key == "mcp_custom"
        else "https"
        if key in web
        else key
    )
    operations = list(definition.discovery_modes)
    if key == "mcp_custom":
        operations = ["validate", "discover_tools", "read", "checkpoint", "close"]
    elif key in database or key in lark or key in web:
        operations = [
            "validate",
            "discover",
            "introspect",
            "sample",
            "read",
            "checkpoint",
            "close",
        ]
    controls = ["secret_ref_only"]
    if key in database:
        controls.extend(
            ["read_only_sql", "parameterized_query", "row_byte_time_limits"]
        )
    if key in web:
        controls.extend(
            ["ssrf_dns_rebinding", "operation_allowlist", "bounded_pagination"]
        )
    if key == "mcp_custom":
        controls.extend(
            [
                "shell_false",
                "secret_ref_runtime_resolution",
                "bounded_timeouts",
                "tool_allowlist",
                "untrusted_output_quarantine",
                "output_budget",
                "process_reap",
            ]
        )
    return AdapterContract(
        connector_key=key,
        protocol=protocol,
        operations=operations,
        configuration_schema=definition.input_schema,
        credential_schema=definition.credential_schema,
        security_controls=controls,
        execution_state=definition.capability_state,
    )


def validate_external_configuration(
    definition: ConnectorDefinition,
    configuration: dict[str, object],
    *,
    web_resolver: Callable[[str], list[str]] | None,
    allow_private_hosts: set[str] | None = None,
) -> None:
    key = definition.connector_key
    database = {
        "postgresql",
        "mysql",
        "oracle",
        "sqlserver",
        "clickhouse",
        "doris",
        "starrocks",
        "snowflake",
        "bigquery",
        "hive",
    }
    web = {"rest_api", "graphql", "web_discovery", "custom_http"}
    if key in database:
        query = str(configuration.get("query") or "SELECT 1")
        parameters = configuration.get("queryParameters", {})
        if not isinstance(parameters, dict):
            raise ValueError("queryParameters must be an object")
        try:
            validate_read_only_sql(
                query,
                parameters={str(name): value for name, value in parameters.items()},
            )
            validate_database_limits(
                row_limit=_integer_option(configuration, "rowLimit", 10_000),
                byte_limit=_integer_option(
                    configuration, "byteLimit", 50 * 1024 * 1024
                ),
                timeout_seconds=_integer_option(configuration, "timeoutSeconds", 30),
            )
        except (TypeError, ValueError) as error:
            raise ValueError(str(error)) from error
    if key in web:
        endpoint = configuration.get("endpoint")
        if isinstance(endpoint, str):
            from .http_transport import validate_network_endpoint

            validate_network_endpoint(
                endpoint,
                resolver=web_resolver,
                allow_private_hosts=allow_private_hosts,
            )
    if key == "mcp_custom":
        transport = configuration.get("transport")
        if transport not in {"stdio", "streamable_http", "sse"}:
            raise ValueError("MCP transport is not supported")
        allowlist = configuration.get("toolAllowlist", [])
        if not isinstance(allowlist, list) or not allowlist:
            raise ValueError("MCP tool allowlist must not be empty")
        validate_mcp_tool(
            str(allowlist[0]),
            allowlist={str(item) for item in allowlist},
            output_bytes=_integer_option(configuration, "outputBytes", 1_000_000),
            max_output_bytes=1_000_000,
        )
        if transport in {"streamable_http", "sse"}:
            endpoint = configuration.get("endpoint")
            if not isinstance(endpoint, str) or not endpoint:
                raise ValueError("remote MCP endpoint is required")
            from .http_transport import validate_network_endpoint

            validate_network_endpoint(
                endpoint,
                resolver=web_resolver,
                allow_private_hosts=allow_private_hosts,
            )


def _integer_option(configuration: dict[str, object], key: str, default: int) -> int:
    value = configuration.get(key, default)
    if isinstance(value, bool) or not isinstance(value, int):
        raise TypeError(f"{key} must be an integer")
    return value
